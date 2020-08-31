    # # Note: To successfully run this application, please download and store chromedriver.exe in the same folder where this program is stored
    # 
    # ## Link to download: https://chromedriver.chromium.org/downloads
    # %% [markdown]


def main(linkedin_job_title, linkedin_job_location, file_name, path) :   
    
    result_string = ['']*3
    # %% [markdown]
    # ## Importing required libraries

    # %%
    from selenium import webdriver
    from selenium.common.exceptions import WebDriverException

    from time import sleep
    from selenium.webdriver.common.keys import Keys
    from parsel import Selector

    from bs4 import BeautifulSoup  
    from urllib.request import urlopen as uReq

    import pandas as pd
    from datetime import date

    from pandas import ExcelWriter
    from openpyxl.utils.dataframe import dataframe_to_rows

    # %% [markdown]
    # ## Specifying path to Chromedriver and initializing driver to https://www.linkedin.com/jobs

    # %%
    def set_chromedriver():
        """
        Args: Takes the path where chromedriver.exe file is stored
        Returns: None
        """

        try:

            # specifies driver
            chromepath = path.replace("\\","/")+'/chromedriver.exe'
            
            head_driver = webdriver.Chrome(chromepath)
            sleep(0.5)

            # driver.get method() will navigate to a page given by the URL address
            head_driver.get('https://www.linkedin.com/jobs')
            sleep(7)

            # Passing head_driver value to job parameter function
            job_parameters(head_driver)

            
            
        except OSError:
            s1 = "Please close excel file and re-run the program"
            result_string[0] = s1
            # quit()
        
        except (OSError,WebDriverException):
            s1 = "**Error: Please check Chromedriver folder path and re-run the program"
            result_string[0] = s1
            # quit()

        return

    # %% [markdown]
    # ## Inputing search parameters from parameters.py for linkedin job search
    # 
    # * Search by title, skill, or company
    # * City, state, or zip code
    #     

    # %%
    def job_parameters(head_driver):

        """
        Mimics keyboard to enter the parameters for job search
        
        Args : None
        Returns: None
        """
        # to clean up the job title textbox
        head_driver.find_element_by_xpath("//input[@placeholder='Search job titles or companies']").clear()

        # locate job title from job_search_inputs
        job_title = head_driver.find_element_by_xpath("//input[@placeholder='Search job titles or companies']")

        # send_keys() to simulate key strokes
        job_title.send_keys(linkedin_job_title)


        # to clean up the location textbox
        head_driver.find_element_by_xpath("//input[@placeholder='Location']").clear()

        # locate job location from job_search_inputs
        job_location = head_driver.find_element_by_xpath("//input[@placeholder='Location']")

        # send_keys() to simulate key strokes
        job_location.send_keys(linkedin_job_location)

        # sleep for 0.5 seconds
        sleep(0.5)


        # locate job search button
        job_location.send_keys(Keys.TAB)
        job_location.send_keys(Keys.TAB)
        job_location.send_keys(Keys.ENTER)

        # sleep for 5 seconds to load the result page
        sleep(5)

        page_scroll(head_driver)

        return


    # %%
    def page_scroll(head_driver):

        """
        This fundtion is required to scroll the search page till the end and make all
        jobs visible
        Args: head driver
        Returns: None
        """

        lenOfPage = head_driver.execute_script("window.scrollTo(0, document.body.scrollHeight);var lenOfPage=document.body.scrollHeight;return lenOfPage;")
        match=False
        while(match==False):
                lastCount = lenOfPage
                sleep(3)
                lenOfPage = head_driver.execute_script("window.scrollTo(0, document.body.scrollHeight);var lenOfPage=document.body.scrollHeight;return lenOfPage;")
                
                if lastCount==lenOfPage:
                    match=True
        while True:
            try:
                head_driver.find_element_by_xpath('/html/body/main/div/section/button').click()
            except:
                # print('All job listings visible')
                break

        sleep(3)
        # job_postings = job_containers(head_driver)
        job_containers(head_driver)
        return 

    # %% [markdown]
    # ## Finding containers on job search page holding individual job postings

    # %%
    def job_containers(head_driver):

        """
        This function will find all the job postings on the pages
        Args: head_driver
        Returns: None
        """

        pageSource = head_driver.page_source
        lxml_soup = BeautifulSoup(pageSource, 'lxml')

        # searching for all job postings
        job_postings = lxml_soup.find('ul', class_ = 'jobs-search__results-list')
        sleep(2)

        # print('Total', len(job_postings), 'job postings found for',linkedin_job_title, 'position in', linkedin_job_location)
        s1 = 'Total ' + str(len(job_postings)) + ' job postings found for ' + linkedin_job_title + ' position in ' + linkedin_job_location
        result_string[0] = s1
        head_driver.quit()

        job_scraping(job_postings, lxml_soup)
        
        return

    # %% [markdown]
    # ## Collecting job details to list from main search page

    # %%
    def job_scraping(job_postings, lxml_soup):
        
        """
        This function will collect details of the job from each container and store them in list
        Args: Job containers
        Return: None
        """

        job_titles = []
        company_names = []
        job_locations = []
        posting_dates = []
        job_links = []

        job_ids = [item['data-id'] for item in lxml_soup.find_all('li', attrs={'data-id' : True})]

        for job in job_postings:
            
            #Scrapping info from main search page
            try:
                job_title = job.find("span", class_="screen-reader-text").text
            except:
                job_title = 'No data available'
            job_titles.append(job_title)

            try:
                company_name = job.select_one('img')['alt']
            except:
                company_name = 'No data available'
            company_names.append(company_name)

            try:
                job_location = job.find("span", class_="job-result-card__location").text
            except:
                job_location = 'No data available'
            job_locations.append(job_location)

            try:
                posting_date = job.select_one('time')['datetime']
            except:
                posting_date = 'No data available'
            posting_dates.append(posting_date)

            try:
                job_link = job.find('a', href=True)['href']
            except:
                job_link = 'No data available'
            job_links.append(job_link)

        job_posting_info = to_df(job_ids, job_titles, company_names, job_locations, posting_dates, job_links)
        check_dublicates(job_posting_info)

        return


    # %%
    #Saving data to dataframe

    def to_df(job_ids, job_titles, company_names, job_locations, posting_dates, job_links):
        
        job_posting_info = pd.DataFrame({'Job Title': job_titles,
                                        'Company Name': company_names,
                                        'Location': job_locations,
                                        'Posting Date': posting_dates,
                                        'Website': job_links,
                                        'Job-ID': job_ids,
                                        })
        return job_posting_info


    # %%
    def check_dublicates(job_posting_info):


        filename = file_name #File name to store the scraped data

        try:
            # Old job postings, already saved in existing excel files will be removed
            # Comparing Job IDs for job postings, we will find the dublicate entries
            old_job_posting_info = pd.read_excel(filename, dtype=str)
            old_job_posting_info.drop(['Level', 'Type', 'Function'], axis=1, inplace=True)

            
            # find jobs in job_posting_info that are not in old_job_posting_info
            new_job_posting_info = job_posting_info[~(job_posting_info['Job-ID'].isin(old_job_posting_info['Job-ID']))].dropna().reset_index(drop=True)
        

            #If no new jobs are added
            if(new_job_posting_info.shape[0] == 0):
                # print('No new jobs found')
                s2 = 'No new jobs found'
                result_string[1] = s2
                return
            
            #If new jobs are added, only those will be saved in new_job_posting_info
            else:
                # print(new_job_posting_info.shape[0], "new jobs found")
                s2 = str(new_job_posting_info.shape[0]) + ' new jobs found'
                result_string[1] = s2
                #Calling job_details function for new jobs to get additional details
                new_job_posting_info = job_details(new_job_posting_info)

                #Writing new job details to a sheet name under today's date
                append_df_to_excel(filename, new_job_posting_info, sheet_name=str(date.today()), startcol=0, startrow=None)      
                #Appending new jobs to master data sheet
                append_df_to_excel(filename, new_job_posting_info, sheet_name="master_job_data", startcol=0, startrow=None, header=False)
                # print(new_job_posting_info.shape[0], "new jobs appended to master job data")
                s3 = str(new_job_posting_info.shape[0]) + " new jobs appended to master job data"
                result_string[2] = s3
                return

        #If we are running code foe first time and there is no master data sheet
        except FileNotFoundError:
            # print(job_posting_info.shape[0], "new jobs found")
            s2 = str(job_posting_info.shape[0]) + " new jobs found"
            result_string[1] = s2
            #Calling job_details function for all jobs to get additional details
            job_posting_info = job_details(job_posting_info)
            
            #Creating a master data sheet for all jobs found
            append_df_to_excel(filename, job_posting_info, sheet_name="master_job_data", startcol=0, startrow=None)
            # print(job_posting_info.shape[0], "jobs saved to master job data")   
            s3 = str(job_posting_info.shape[0]) + " jobs saved to master job data"  
            result_string[2] = s3
            return


    # %%
    #scrapping from individual job posting
    def job_details(job_posting_info):

        """
        This function will use the web link of job posting to open and collect additional information fo
        the job.

        Args: job posting dataframe
        
        Returns: job posting dataframe with additional columns 
                    List of job seniority level, 
                    List of job types, 
                    List of job functions.
        """

        job_seniority_levels = []
        job_types = []
        job_functions = []

        job_links = job_posting_info['Website']

        for job_link in job_links:
            
            job_page_client  = uReq(job_link) #Grabbing the webpage stored in mu_url
            job_page_html    = job_page_client.read() 
            sleep(1)
            job_page_client.close() #Close the web client after grabbing the data
            soup = BeautifulSoup(job_page_html, "html.parser") #Parsing the file in html format
            job_details = soup.findAll('span',{'class':['job-criteria__text job-criteria__text--criteria']})

            try:
                job_seniority_level = job_details[0].text
            except:
                job_seniority_level = 'No data Available'
            job_seniority_levels.append(job_seniority_level)

            try:
                job_type = job_details[1].text
            except:
                job_type = 'No data Available'
            job_types.append(job_type)

            try:
                job_function = job_details[2].text
            except:
                job_function = 'No data Available'
            job_functions.append(job_function)


        #Add new columns to dataframe
        job_posting_info['Level'] = job_seniority_levels
        job_posting_info['Type']  = job_types
        job_posting_info['Function'] = job_functions

        return(job_posting_info)


    # %%
    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                        truncate_sheet=False, 
                        **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
        filename : File path or existing ExcelWriter
                    (Example: '/path/to/file.xlsx')
        df : dataframe to save to workbook
        sheet_name : Name of sheet which will contain DataFrame.
                    (default: 'Sheet1')
        startrow : upper left cell row to dump data frame.
                    Per default (startrow=None) calculate the last row
                    in the existing DF and write to the next row...
        truncate_sheet : truncate (remove and recreate) [sheet_name]
                        before writing DataFrame to Excel file
        to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """
        try:

            from openpyxl import load_workbook

            # ignore [engine] parameter if it was passed
            if 'engine' in to_excel_kwargs:
                to_excel_kwargs.pop('engine')

            writer = pd.ExcelWriter(filename, engine='openpyxl') # pylint: disable=abstract-class-instantiated

            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs, index=False)

        # save the workbook
        writer.save()


    # %%
    set_chromedriver()

    return (result_string)
